import pandas as pd
import numpy as np
import os
import json
import anthropic
from datetime import datetime
import fitz  # PyMuPDF pour lire les PDF
import openpyxl
from io import BytesIO
import asyncio

CLAUDE_MODEL = "claude-sonnet-4-20250514"

# Fonction asynchrone pour le Chatbot (utilisée par main.py)
async def ask_claude_general(prompt, api_key):
    if not api_key:
        return "⚠️ Clé API manquante pour Claude."

    client = anthropic.Anthropic(api_key=api_key)

    try:
        # Appel Claude (async)
        response = await asyncio.to_thread(
            lambda: client.messages.create(
                model=CLAUDE_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1000,
                temperature=0.7
            )
        )
        return response.content[0].text

    except Exception as e:
        return f"⚠️ Erreur Claude : {str(e)}"

class AuditEngine:
    def __init__(self, mission_id):
        self.mission_id = mission_id
        # Récupération sécurisée de la clé
        self.api_key = os.environ.get("ANTHROPIC_API_KEY")
        
        if self.api_key:
            self.claude_client = anthropic.Anthropic(api_key=self.api_key)
        else:
            self.claude_client = None
            print("⚠️ ATTENTION: Clé API Anthropic manquante !")

    def lire_fichier_universel(self, contents, filename):
        """Lit PDF, Excel, Word ou TXT/CSV et renvoie du texte"""
        try:
            if filename.endswith('.pdf'):
                doc = fitz.open(stream=contents, filetype="pdf")
                texte = ""
                for page in doc:
                    texte += page.get_text() + "\n"
                return texte

            elif filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
                texte = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        line = " ".join([str(c) for c in row if c is not None])
                        texte += line + "\n"
                return texte

            # --- AJOUT DU SUPPORT WORD ---
            elif filename.endswith('.docx'):
                from docx import Document
                doc = Document(BytesIO(contents))
                return "\n".join([para.text for para in doc.paragraphs])
            # -----------------------------

            else:
                return contents.decode("utf-8", errors="ignore")
        except Exception as e:
            return f"ERREUR LECTURE FICHIER: {str(e)}"
        

        
    def _determiner_cycle(self, compte_num):
        """Assigne un des 21 cycles d'audit"""
        c = str(compte_num).strip()
        if c.startswith(('10', '11', '12')): return "CAPITAUX_PROPRES"
        if c.startswith('15'): return "PROVISIONS"
        if c.startswith('16'): return "EMPRUNTS"
        if c.startswith('2'): return "IMMO_CORPORELLES" 
        if c.startswith('3'): return "STOCKS"
        if c.startswith('40'): return "FOURNISSEURS"
        if c.startswith('41'): return "CLIENTS"
        if c.startswith(('42', '43')): return "DETTES_SOCIALES"
        if c.startswith('44'): return "DETTES_FISCALES"
        if c.startswith(('45', '46')): return "AUTRES_CREANCES_DETTES"
        if c.startswith('5'): return "TRESORERIE"
        if c.startswith(('60', '61', '62')): return "CHARGES_EXPLOITATION"
        if c.startswith('64'): return "CHARGES_PERSONNEL"
        if c.startswith('66'): return "RESULTAT_FINANCIER"
        if c.startswith('69'): return "IMPOTS"
        if c.startswith('70'): return "PRODUITS_EXPLOITATION"
        return "OPERATIONS_DIVERSES"

    def executer_analyse_v4(self, df):
        """Analyse principale du FEC"""
        anomalies = []
        
        # Nettoyage
        for col in ['debit', 'credit']:
            if df[col].dtype == object:
                df[col] = df[col].astype(str).str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        if 'compte_num' in df.columns:
            df['cycle_calcule'] = df['compte_num'].apply(self._determiner_cycle)
        else:
            df['cycle_calcule'] = "INCONNU"

        # 1. Analyse Python (Règles strictes)
        try:
            anomalies.extend(self._analyse_benford(df))
            anomalies.extend(self._analyse_tracfin(df))
            anomalies.extend(self._analyse_comptes_sensibles(df))
        except Exception as e:
            print(f"Erreur analyse Python : {e}")
        
        # 2. Analyse IA (Claude)
        try:
            # Filtre des écritures suspectes pour l'IA
            df['is_round'] = (df['debit'] % 100 == 0) & (df['debit'] > 1000)
            keywords = ['divers', 'regularisation', 'gift', 'cadeau', 'espece', 'honoraires', 'consulting']
            df['is_suspect'] = df['ecriture_lib'].astype(str).str.contains('|'.join(keywords), case=False, na=False)
            
            risky_lines = df[
                df['is_round'] | df['is_suspect'] | (df['journal_code'] == 'OD')
            ].sort_values(by='debit', ascending=False).head(30)
            
            # Si pas assez de suspect, on prend les plus gros montants
            if len(risky_lines) < 5:
                risky_lines = df.sort_values(by='debit', ascending=False).head(10)

            cols = ['journal_code', 'ecriture_date', 'compte_num', 'ecriture_lib', 'debit', 'credit', 'cycle_calcule']
            final_cols = [c for c in cols if c in risky_lines.columns]
            data_json = risky_lines[final_cols].to_json(orient="records")

            if self.api_key:
                ai_anomalies = self._ask_claude_expert(data_json)
                anomalies.extend(ai_anomalies)

        except Exception as e:
            print(f"Erreur IA : {e}")

        return anomalies
    


    # --- FONCTION MODIFIÉE (REMPLACE L'ANCIENNE) ---
    def analyser_document_pdf_excel(self, texte, is_sigle=False, cycle_id=None): # <--- AJOUT cycle_id
        """Analyse IA avec détection d'écarts (SIGLES)"""
        import re
        import json

        # Nettoyage texte
        texte_propre = re.sub(r'[^\x20-\x7E\nÀ-ÿ€]', '', texte)

        # Consigne par défaut
        consigne_audit = "Cherche des incohérences de dates, montants suspects, mentions de litiges."

        # AJOUT : Instruction spécifique pour le code C01, C02...
        instruction_code = ""
        if cycle_id:
            instruction_code = f"IMPORTANT : Pour le champ 'cycle', utilise impérativement le code '{cycle_id}'."
        else:
            instruction_code = "Classe chaque anomalie dans l'un de ces CYCLES : IMMO_CORPORELLES, STOCKS, CLIENTS, TRESORERIE, etc."

        # Consigne spéciale SIGLES (Ta logique conservée et enrichie)
        if is_sigle:
            consigne_audit = f"""
    FOCUS AUDIT (SIGLES) :
    {f"Tu analyses actuellement le cycle spécifique : {cycle_id}" if cycle_id else ""}
    1. Analyse ce document comme un bilan (Actif ou Passif).
    2. Identifie les variations atypiques entre N et N-1.
    3. Vérifie la cohérence des rubriques (ex: amortissements vs immo).
    4. Si tu détectes un écart de conversion ou de solde, signale-le comme CRITIQUE.

    TEXTE :
    {texte_propre[:20000]}

    RÈGLE DE CLASSEMENT :
    {instruction_code}

    Réponds UNIQUEMENT en JSON Array :
    [
    {{
        "cycle": "{cycle_id if cycle_id else 'TRESORERIE'}",
        "type_anomalie": "DOCUMENTAIRE",
        "niveau_criticite": "ELEVE",
        "score_ml": 80,
        "montant": 15000,
        "description": "Explication claire..."
    }}
    ]
    Si rien à signaler, renvoie [].
    """

        prompt = consigne_audit

        try:
            message = self.claude_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=2000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            content = message.content[0].text
            start = content.find('[')
            end = content.rfind(']') + 1
            if start == -1 or end == 0: return []
            return json.loads(content[start:end])
        except Exception as e:
            print("Erreur analyse IA :", e)
            return []


    # --- NOUVELLE FONCTION A AJOUTER ---
    def convertir_excel_en_df(self, contents):
        """Tente de convertir un Excel en DataFrame structuré pour l'analyse chiffrée"""
        try:
            # On lit le Excel avec Pandas direct
            df = pd.read_excel(BytesIO(contents))
            
            # Normalisation des colonnes (minuscules + strip)
            df.columns = [str(c).strip().lower() for c in df.columns]
            
            # Dictionnaire de synonymes pour retrouver les colonnes FEC
            mapping = {
                'compte': 'compte_num', 'num_compte': 'compte_num', 'n° compte': 'compte_num',
                'libellé': 'ecriture_lib', 'libelle': 'ecriture_lib', 'label': 'ecriture_lib',
                'débit': 'debit', 'debit': 'debit',
                'crédit': 'credit', 'credit': 'credit',
                'journal': 'journal_code', 'code': 'journal_code',
                'date': 'ecriture_date'
            }
            df.rename(columns=mapping, inplace=True)
            
            # Si on a au moins Débit ou Crédit, on considère que c'est exploitable en chiffres
            if 'debit' in df.columns or 'credit' in df.columns:
                # Ajout des colonnes manquantes par défaut pour éviter les bugs
                for col in ['debit', 'credit']: 
                    if col not in df.columns: df[col] = 0
                if 'compte_num' not in df.columns: df['compte_num'] = '000000'
                if 'ecriture_lib' not in df.columns: df['ecriture_lib'] = 'Import Excel'
                if 'journal_code' not in df.columns: df['journal_code'] = 'OD'
                
                return df
            return None
        except:
            return None

        # AJOUT POUR LA CIRCULARISATION
    async def extraire_tiers_pour_circularisation(self, texte_donnees, type_tiers):
        prompt = f"""
        Tu es un assistant audit expert. Voici le contenu brut d'une Balance Auxiliaire {type_tiers}.
        
        CONTENU DU FICHIER :
        {texte_donnees[:20000]} 

        TA MISSION :
        1. Identifie chaque Tiers (Nom de l'entreprise/client).
        2. Identifie le Solde comptable associé.
        3. Si tu trouves une adresse email, prends-la, sinon laisse vide.
        
        FORMAT DE RÉPONSE ATTENDU (JSON UNIQUEMENT, ARRAY) :
        [
            {{ "nom": "CLIENT SARL", "montant": "1500.00", "email": "contact@client.com" }},
            {{ "nom": "DUBOSC SA", "montant": "4200.50", "email": "" }}
        ]
        
        Ne renvoie que du JSON valide. Ignore les totaux généraux, prends les lignes individuelles.
        """
        try:
            if not self.claude_client: return []
            
            # OPTIMISATION : On utilise asyncio pour ne pas bloquer le serveur
            response = await asyncio.to_thread(
                lambda: self.claude_client.messages.create(
                    model=CLAUDE_MODEL, 
                    max_tokens=3000,
                    temperature=0,
                    messages=[{"role": "user", "content": prompt}]
                )
            )
            content = response.content[0].text
            
            # Parsing JSON
            start = content.find('[')
            end = content.rfind(']') + 1
            return json.loads(content[start:end])
        except Exception as e:
            print(f"Erreur extraction tiers: {e}")
            return []


    def _ask_claude_expert(self, json_data):
        prompt = f"""
        AGIS COMME UN AUDITEUR. Analyse ces écritures (JSON).
        Règles : Cherche fraudes, doublons, erreurs fiscales.
        DONNÉES : {json_data}
        Réponds UNIQUEMENT en JSON : 
        [{{ "cycle": "...", "type_anomalie": "...", "niveau_criticite": "CRITIQUE/ELEVE", "score_ml": 90, "montant": 0, "description": "..." }}]
        """
        try:
            message = self.claude_client.messages.create(
                model=CLAUDE_MODEL, 
                max_tokens=2500,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            content = message.content[0].text
            start = content.find('[')
            end = content.rfind(']') + 1
            return json.loads(content[start:end])
        except: return []

    def _analyse_benford(self, df):
        target_df = df[df['compte_num'].astype(str).str.startswith(('6', '7'))]
        if len(target_df) < 50: return []
        first_digits = target_df['debit'].astype(str).str[:1]
        first_digits = first_digits[first_digits.str.isnumeric()].astype(int)
        freq_1 = (first_digits == 1).mean()
        if freq_1 < 0.20 or freq_1 > 0.40:
            return [{
                "cycle": "CHARGES_EXPLOITATION", 
                "type_anomalie": "STATISTIQUE (BENFORD)",
                "niveau_criticite": "ELEVE",
                "score_ml": 85.0, 
                "montant": 0,
                "description": f"Anomalie Benford : Fréquence du '1' anormale ({round(freq_1*100, 1)}%)."
            }]
        return []

    def _analyse_tracfin(self, df):
        treso = df[df['compte_num'].astype(str).str.startswith('5')]
        smurfing = treso[(treso['debit'] >= 9000) & (treso['debit'] < 10000)]
        if len(smurfing) >= 1:
            return [{
                "cycle": "TRESORERIE", 
                "type_anomalie": "BLANCHIMENT (TRACFIN)",
                "niveau_criticite": "CRITIQUE",
                "score_ml": 98.0, 
                "montant": float(smurfing['debit'].sum()),
                "description": f"ALERTE : {len(smurfing)} mvts entre 9k€ et 10k€ (soupçon fractionnement)."
            }]
        return []

    def _analyse_comptes_sensibles(self, df):
        if 'compte_num' in df.columns:
            caisse = df[df['compte_num'].astype(str).str.startswith('53')]
            if not caisse.empty:
                solde = caisse['debit'].sum() - caisse['credit'].sum()
                if solde < -10:
                     return [{
                        "cycle": "TRESORERIE",
                        "type_anomalie": "COHÉRENCE COMPTABLE",
                        "niveau_criticite": "CRITIQUE",
                        "score_ml": 100.0, 
                        "montant": float(abs(solde)),
                        "description": f"Caisse négative de {abs(solde)} €. Impossible physiquement."
                    }]
        return []
